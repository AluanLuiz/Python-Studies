# Atividade prática python/excel
# Aluan Luiz Maturi Horshcutz - 2323033
# Eduardo Rufo Bruscagin - 2323008

#--- Imports
import tkinter as tk
from tkinter import filedialog
import random as rdm
import os
from openpyxl import Workbook
from openpyxl import load_workbook

#-- Fontes
high=("Arial", 22)
medium=("Arial", 17)
small=("Arial", 16)

#--- Interface Menu
class Menu:
    def __init__(self, master):
        self.main = master
        self.main.title("Campo Minado - P4")
        self.main.configure(bg="#B6B6B6")
        
        self.frm_campo = tk.Frame(self.main, bg="#B6B6B6")
        self.frm_campo.grid(row=1, column=0, columnspan=5, sticky="ew")
        
        self.frm_btns = tk.Frame(self.main, bg="#B6B6B6")
        self.frm_btns.grid(row=4, column=2, sticky="ews")
        
        self.main.grid_rowconfigure(5, weight=1)
        self.main.grid_columnconfigure((0, 1, 2, 3, 4), weight=1)

        self.frm_campo.grid_rowconfigure(4, weight=1)
        self.frm_campo.grid_columnconfigure((0, 1, 2, 3, 4), weight=1)

        self.frm_btns.grid_rowconfigure(3, weight=1)
        self.frm_btns.grid_columnconfigure((0,1,2), weight=1)
        
        self.entrys()
        self.buttons()
        
    def entrys(self):
        #-- Texto de Titulo
        self.lbl_title = tk.Label(self.frm_campo, font=high, text="Minefield \n Campo Minado", bg="#B6B6B6")
        self.lbl_title.grid(row=1, column=1, columnspan=3, padx=5, pady=20, sticky="new")
        
    def buttons(self):
        #-- ↓ botao jogar
        self.btn_play = tk.Button(self.frm_btns, font=medium, text="Play", 
                                  width=10, height=2, bd=4, highlightthickness=1, bg="#D0D0D0", command=self.play_command)
        self.btn_play.grid(row=0, column=1, padx=5, pady=5)
        
        #-- ↓ botao configurar
        self.btn_config = tk.Button(self.frm_btns, font=medium, text="Settings", 
                                    width=10, height=2, bd=4, highlightthickness=1, bg="#D0D0D0", command= self.config_command)
        self.btn_config.grid(row=1, column=1, padx=5, pady=5)

        #-- ↓ botao sair
        self.btn_exit = tk.Button(self.frm_btns, font=medium, text="Exit", 
                                  width=10, height=2, bd=4, highlightthickness=1, bg="#D0D0D0", command=exit)
        self.btn_exit.grid(row=2, column=1, padx=5, pady=5)
    
    def config_command(self): #Comando para abrir a tela de configuracao
        self.clear_window()
        Config(self.main)

    def clear_window(self):
        for widget in self.main.winfo_children():
            widget.destroy()

    def play_command(self): # Comando para iniciar o jogo em outra tela
        start = tk.Toplevel(self.main)
        StartGame(start)

    def exit(self):
        self.main.destroy()

#--- Interface Config
class Config:
    def __init__(self, setting):
        self.settings = setting
        self.settings.title("Configuração do Jogo")
        self.settings.configure(bg="#B6B6B6")
        self.settings.geometry("400x400")
        
        self.difficulty_var = tk.StringVar(value="Fácil")
        
        self.frm_campo2 = tk.Frame(self.settings, bg="#B6B6B6")
        self.frm_campo2.grid(row=1, column=0, columnspan=5, sticky="ew")
        
        self.frm_btns2 = tk.Frame(self.settings, bg="#B6B6B6")
        self.frm_btns2.grid(row=4, column=2, sticky="ews")
        
        self.settings.grid_rowconfigure(5, weight=1)
        self.settings.grid_columnconfigure((0, 1, 2, 3, 4), weight=1)

        self.frm_campo2.grid_rowconfigure(5, weight=1)
        self.frm_campo2.grid_columnconfigure((0, 1, 2, 3), weight=1)

        self.frm_btns2.grid_rowconfigure(4, weight=1)
        self.frm_btns2.grid_columnconfigure((0,1,2), weight=1)
        
        self.entrys2()
        self.buttons2()
        
    def entrys2(self):
        #--- ↓ texto titulo
        self.lbl_title_2 = tk.Label(self.frm_campo2, font=high, text="Configuração", bg="#B6B6B6")
        self.lbl_title_2.grid(row=0, column=1, columnspan=3, padx=5, pady=20, sticky="new")
        
        #--- ↓ Quantidade de Linhas
        self.lbl_qtdlines = tk.Label(self.frm_campo2, font=small, text="Definir qtd de Linhas: ", bg="#B6B6B6")
        self.spn_qtdLines = tk.Spinbox(self.frm_campo2, font=small, from_=5, to= 50, width=10)

        self.lbl_qtdlines.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        self.spn_qtdLines.grid(row=1, column=2, padx=5, pady=5, sticky="e")
        
        #--- ↓ Quantidade de Colunas
        self.lbl_qtdcolumn = tk.Label(self.frm_campo2, font=small, text="Definir qtd de Colunas: ", bg="#B6B6B6")
        self.spn_qtdcolumn = tk.Spinbox(self.frm_campo2, font=small, from_=5, to= 50, width=10)

        self.lbl_qtdcolumn.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        self.spn_qtdcolumn.grid(row=2, column=2, padx=5, pady=5, sticky="e")
               
        #----- ↓ Definir dificuldade
        self.lbl_diffic = tk.Label(self.frm_campo2, font=medium, text="Dificuldade:", bg="#B6B6B6")
        self.lbl_diffic.grid(row=3, column=1, columnspan=2 ,padx=5, pady=5, sticky="ew")
        #-----
        
    def buttons2(self):
        #- ↓ RadioButton Facil
        self.rbtn_easy = tk.Radiobutton(self.frm_btns2, text="Fácil", variable=self.difficulty_var, value="Fácil", font=medium, bg="#B2B3B3")
        self.rbtn_easy.grid(row=0, column=0, sticky="w")
        #- ↓ RadioButton Medio
        self.rbtn_medium = tk.Radiobutton(self.frm_btns2, text="Médio", variable=self.difficulty_var, value="Médio", font=medium, bg="#B2B3B3")
        self.rbtn_medium.grid(row=0, column=1)
        #- ↓ RadioButton Dificil
        self.rbtn_hard = tk.Radiobutton(self.frm_btns2, text="Difícil", variable=self.difficulty_var, value="Difícil", font=medium, bg="#B2B3B3")
        self.rbtn_hard.grid(row=0, column=2, sticky="e")
        
        #- ↓ Button Salvar
        self.btn_save = tk.Button(self.frm_btns2, font=small, text="Salvar", 
                                  width=8, height=1, bd=4, highlightthickness=1, bg="#D0D0D0", command=self.save_config)
        self.btn_save.grid(row=2, column=1, padx=5, pady=24, sticky="ew")
        
        #- ↓ Button Voltar
        self.btn_voltar = tk.Button(self.frm_btns2, font=small, text="Voltar", 
                                  width=8, height=1, bd=4, highlightthickness=1, bg="#D0D0D0", command=self.return_)
        self.btn_voltar.grid(row=3, column=1, padx=5, pady=10, sticky="ew")
    
    def return_(self):
        self.clear_window()
        Menu(self.settings)
        
    def clear_window(self):
        for widget in self.settings.winfo_children():
            widget.destroy()
    
    def save_config(self):
        qtd_lines = int(self.spn_qtdLines.get()) #- Extrai o valor inteiro do Spinbox linhas
        qtd_columns = int(self.spn_qtdcolumn.get()) #- Extrai o valor inteiro do Spinbox colunas
        diffi = self.difficulty_var.get() #- Extrai a variavel selecionada do RadioButton
        
        #-- Difinindo a qtd de minas de acordo com a dificuldade
        if diffi == "Fácil": # 20%
            mine = qtd_lines * qtd_columns * 0.2
        elif diffi == "Médio": # 40%
            mine = qtd_lines * qtd_columns * 0.4
        elif diffi == "Difícil": # 60%
            mine = qtd_lines * qtd_columns * 0.6
        else:
            Menu.main.destroy()

        mines = round(mine) #- Arredondando a qtd de bombas
        self.spaw_mines(qtd_lines, qtd_columns, mines)
        self.spaw_numbers(qtd_lines, qtd_columns)
    
        print(f"dimensao: {qtd_lines}x{qtd_columns} ({qtd_lines*qtd_columns}) | dificuldade: {diffi} | bombas: {mines}") #- Teste de extracao de valores

        self.return_()

    def spaw_mines(self, line, colum, mine): #- Criando o xlsx e colocando as bombas/minas no tabuleiro(# = bomba)
        arquivo_1 = Workbook()
        aba_1 = arquivo_1.active
        aba_1.title="tabuleiro"
        
        mines_created = 0
        while mines_created < mine:
            x =  rdm.randint(1, colum)
            y = rdm.randint(1, line)
            if aba_1.cell(row=x, column=y).value != "#":
                aba_1.cell(row=x, column=y, value="#")
                mines_created += 1
        arquivo_1.save("minefield_v1.xlsx")
        
    def spaw_numbers(self, linha, coluna): #- Colocando os numeros no tabuleiro
        open_arquivo = load_workbook(filename="minefield_v1.xlsx", read_only=False)
        aba_ativa = open_arquivo.active
        
        for x in range(1, linha+1):
            for y in range(1, coluna+1):
                cell_value = aba_ativa.cell(row=x, column=y).value
                if cell_value != "#":
                    mines = 0
                    for line in range(max(1,x-1), min(linha, x+1) + 1):
                        for column in range(max(1, y-1), min(coluna, y+1) + 1):
                            cell_value2 = aba_ativa.cell(row=line, column=column).value
                            if cell_value2 == "#":
                                mines +=1
                    aba_ativa.cell(row=x, column=y, value=mines)
                    
        aba_2=open_arquivo.create_sheet("jogo")
        
        # for x in range(2, linha+2):
        #     cel=f"L{x-1}"
        #     aba_2.cell(row=x, column=1,value=cel)

        # for y in range(2,coluna+2):
        #     cel=f"C{y-1}"
        #     aba_2.cell(row=1, column=y,value=cel)

        for x in range(1,linha+1):
            for y in range(1,coluna+1):
                aba_2.cell(row=x, column=y, value="_")
                
        open_arquivo.save("minefield_v1.xlsx")
        open_arquivo.close()

#--- Interface Jogo
class StartGame:
    def __init__(self, game):
        self.game = game
        self.game.title("Jogo")
        self.game.configure(bg="#B6B6B6")
        self.game.geometry("600x600")

        self.frm_widget = tk.Frame(self.game, bg="#B6B6B6")
        self.frm_widget.grid(row=0, column=0, columnspan=5, sticky="ew", padx=10, pady=10)

        self.frm_btns3 = tk.Frame(self.game, bg="#B6B6B6")
        self.frm_btns3.grid(row=1, column=0, columnspan=5, sticky="ews")

        self.game.grid_rowconfigure(2, weight=1)
        self.game.grid_columnconfigure((0, 1, 2, 3, 4), weight=1)

        self.frm_widget.grid_rowconfigure(0, weight=1)
        self.frm_widget.grid_columnconfigure((0, 1, 2, 3, 4), weight=1)

        self.frm_btns3.grid_rowconfigure(1, weight=1)
        self.frm_btns3.grid_columnconfigure((0, 1, 2), weight=1)

        self.load_board()
        self.create_board_ui()
        self.create_widgets()

    def load_board(self):
        self.workbook = load_workbook("minefield_v1.xlsx")
        self.sheet = self.workbook["jogo"]
        self.solution_sheet = self.workbook["tabuleiro"]
        self.board = [[self.sheet.cell(row=i, column=j).value 
                        for j in range(1, self.sheet.max_column + 1)] 
                            for i in range(1, self.sheet.max_row + 1)]
        
        self.mine_locations = [(i, j) for i in range(len(self.board)) 
                               for j in range(len(self.board[0])) 
                                    if self.solution_sheet.cell(row=i+1, column=j+1).value == "#"]
        self.flags = set()
        self.revealed = set()

    def create_board_ui(self):
        self.board_frame = tk.Frame(self.frm_widget, bg="#B6B6B6")
        self.board_frame.grid(row=0, column=0, columnspan=5, padx=10, pady=10)

        self.labels = []
        for i, row in enumerate(self.board):
            row_labels = []
            for j, cell in enumerate(row):
                label = tk.Label(self.board_frame, text=cell if cell != 0 else "", width=4, height=2, font=small, relief="raised", bg="#D0D0D0")
                label.grid(row=i, column=j, padx=1, pady=1)
                row_labels.append(label)
            self.labels.append(row_labels)

    def create_widgets(self):
        self.lbl_line = tk.Label(self.frm_btns3, text="Linha:", font=small, bg="#B6B6B6")
        self.lbl_line.grid(row=0, column=0, padx=5, pady=5)

        self.entry_line = tk.Entry(self.frm_btns3, font=small)
        self.entry_line.grid(row=0, column=1, padx=5, pady=5)

        self.lbl_column = tk.Label(self.frm_btns3, text="Coluna:", font=small, bg="#B6B6B6")
        self.lbl_column.grid(row=1, column=0, padx=5, pady=5)

        self.entry_column = tk.Entry(self.frm_btns3, font=small)
        self.entry_column.grid(row=1, column=1, padx=5, pady=5)

        self.var_is_mine = tk.BooleanVar()
        self.chk_is_mine = tk.Checkbutton(self.frm_btns3, text="Marcar como Bomba", variable=self.var_is_mine, font=small, bg="#B6B6B6")
        self.chk_is_mine.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

        self.btn_submit = tk.Button(self.frm_btns3, text="Salvar", font=small, command=self.save_move)
        self.btn_submit.grid(row=3, column=0, columnspan=2, padx=5, pady=5)

    def save_move(self):
        try:
            line = int(self.entry_line.get()) - 1
            column = int(self.entry_column.get()) - 1
            is_mine = self.var_is_mine.get()
        except ValueError:
            print("Por favor, insira valores válidos.")
            return

        if line < 0 or line >= len(self.board) or column < 0 or column >= len(self.board[0]):
            print("Posição fora do tabuleiro.")
            return

        if is_mine:
            if (line, column) in self.mine_locations:
                self.board[line][column] = "B"
            else:
                self.board[line][column] = "X"  # marcação incorreta de bomba

            self.flags.add((line, column))

        else:
            value_from_solution = self.solution_sheet.cell(row=line+1, column=column+1).value
            self.board[line][column] = value_from_solution
            self.revealed.add((line, column))

            if value_from_solution == 0:
                self.reveal_empty_cells(line, column)

        self.update_board()

        if self.check_win():
            print("Você ganhou!")
            self.show_all_bombs()
            return

        if self.check_loss(line, column):
            print("Você perdeu!")
            self.show_all_bombs()
            return

        print(f"Linha: {line+1}, Coluna: {column+1}, É bomba: {is_mine}")

    def update_board(self):
        for i in range(len(self.board)):
            for j in range(len(self.board[0])):
                text = self.board[i][j] if self.board[i][j] != 0 else ""
                self.labels[i][j].config(text=text)

    def reveal_empty_cells(self, row, col):
        stack = [(row, col)]
        visited = set(stack)

        while stack:
            r, c = stack.pop()
            for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1), (-1, -1), (-1, 1), (1, -1), (1, 1)]:
                nr, nc = r + dr, c + dc
                if 0 <= nr < len(self.board) and 0 <= nc < len(self.board[0]) and (nr, nc) not in visited:
                    visited.add((nr, nc))
                    if self.board[nr][nc] == 0:
                        stack.append((nr, nc))
                    self.board[nr][nc] = self.solution_sheet.cell(row=nr+1, column=nc+1).value
                    self.revealed.add((nr, nc))

    def check_win(self):
        non_mine_cells = len(self.board) * len(self.board[0]) - len(self.mine_locations)
        return self.flags == set(self.mine_locations) and len(self.revealed) == non_mine_cells

    def check_loss(self, row, col):
        return (row, col) in self.mine_locations

    def show_all_bombs(self):
        for r, c in self.mine_locations:
            self.board[r][c] = "B"
        self.update_board()

                
#----- Iniciar o Menu
def init_minefield():
    minado = tk.Tk()
    app = Menu(minado)
    minado.geometry("400x400")
    minado.mainloop()
    
init_minefield()